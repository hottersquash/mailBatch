"""批量邮件发送核心模块。"""

from __future__ import annotations

import logging
import re
import smtplib
import time
from dataclasses import dataclass, field
from email.header import Header
from email.mime.application import MIMEApplication
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

EMAIL_COLUMN_NAMES = (
    "邮箱",
    "邮件",
    "Email",
    "email",
    "E-mail",
    "e-mail",
    "电子邮箱",
    "邮件地址",
    "邮箱地址",
    "收件邮箱",
    "收件人邮箱",
    "收件邮件",
    "收件人邮件",
    "收件人",
)
CC_COLUMN_NAMES = (
    "抄送",
    "抄送人",
    "抄送邮箱",
    "抄送邮件",
    "抄送人邮箱",
    "抄送人邮件",
    "CC",
    "cc",
    "Cc",
    "Carbon Copy",
)
PLACEHOLDER_PATTERN = re.compile(r"\{\{([^}]+)\}\}")
LEGACY_PLACEHOLDER_PATTERN = re.compile(r"【([^】]+)】")
EMAIL_FORMAT_PATTERN = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")


@dataclass
class SmtpConfig:
    host: str
    port: int
    user: str
    password: str
    use_tls: bool = True
    use_ssl: bool = False
    from_email: str = ""
    from_name: str = ""
    send_interval: float = 1.0

    def __post_init__(self) -> None:
        if not self.from_email:
            self.from_email = self.user

    @property
    def mail_from(self) -> str:
        return self.from_email

    @mail_from.setter
    def mail_from(self, value: str) -> None:
        self.from_email = value


@dataclass
class ExcelData:
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass
class ValidationIssue:
    level: str
    message: str


@dataclass
class SendResult:
    row_index: int
    recipient: str
    success: bool
    error: str = ""
    cc: str = ""


@dataclass
class EmailPreview:
    row_index: int
    to_email: str
    cc_emails: list[str]
    subject: str
    body: str
    attachment_paths: list[str]


SendRowResult = SendResult


@dataclass
class BatchSendSummary:
    total: int = 0
    success_count: int = 0
    failure_count: int = 0
    skipped_count: int = 0
    results: list[SendResult] = field(default_factory=list)

    @property
    def success(self) -> int:
        return self.success_count

    @property
    def failed(self) -> int:
        return self.failure_count


def _parse_bool(value: str | None, default: bool) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def load_smtp_config(
    env_path: str | Path | None = None,
    *,
    host: str | None = None,
    port: int | None = None,
    user: str | None = None,
    password: str | None = None,
    use_tls: bool | None = None,
    use_ssl: bool | None = None,
    from_email: str | None = None,
    from_name: str | None = None,
) -> SmtpConfig:
    """从 .env 和显式参数加载 SMTP 配置，显式参数优先。"""
    env_file = Path(env_path) if env_path else Path(".env")
    if load_dotenv is not None and env_file.is_file():
        load_dotenv(env_file)

    import os

    def pick(name: str, override: str | int | bool | None, default: str = "") -> str:
        if override is not None:
            return str(override)
        return os.getenv(name, default)

    resolved_port = port
    if resolved_port is None:
        resolved_port = int(os.getenv("SMTP_PORT", "587"))

    resolved_use_ssl = use_ssl
    if resolved_use_ssl is None:
        resolved_use_ssl = _parse_bool(os.getenv("SMTP_USE_SSL"), False)

    resolved_use_tls = use_tls
    if resolved_use_tls is None:
        resolved_use_tls = _parse_bool(os.getenv("SMTP_USE_TLS"), not resolved_use_ssl)

    send_interval = float(os.getenv("SEND_INTERVAL", "1.0"))

    return SmtpConfig(
        host=pick("SMTP_HOST", host),
        port=resolved_port,
        user=pick("SMTP_USER", user),
        password=pick("SMTP_PASSWORD", password),
        use_tls=resolved_use_tls,
        use_ssl=resolved_use_ssl,
        from_email=pick("MAIL_FROM", from_email) or pick("SMTP_FROM", from_email),
        from_name=pick("MAIL_FROM_NAME", from_name) or pick("SMTP_FROM_NAME", from_name),
        send_interval=send_interval,
    )


def load_excel(path: str | Path) -> ExcelData:
    """读取 Excel 第一个 sheet。"""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return ExcelData(sheet.title, [], [])

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            if values is None or all(cell is None or str(cell).strip() == "" for cell in values):
                continue
            row_dict: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                row_dict[header] = "" if value is None else value
            rows.append(row_dict)
        return ExcelData(sheet.title, headers, rows)
    finally:
        workbook.close()


def validate_send_request(
    excel_data: ExcelData,
    template_text: str,
    subject: str,
    attachment_paths: Iterable[str | Path],
    *,
    email_column: str | None = None,
    cc_column: str | None = None,
) -> list[ValidationIssue]:
    """发送前校验必填项，不执行 dry-run 预览。"""
    issues: list[ValidationIssue] = []

    if not excel_data.rows:
        issues.append(ValidationIssue("error", "Excel 中没有可发送的数据行。"))

    if not subject.strip():
        issues.append(ValidationIssue("error", "请填写邮件主题。"))

    if not template_text.strip():
        issues.append(ValidationIssue("error", "邮件模板不能为空。"))

    recipient_column = email_column or find_email_column(excel_data.headers)
    if not recipient_column:
        issues.append(
            ValidationIssue(
                "error",
                "未找到收件邮箱列，请在左侧选择「收件人列」，或 Excel 表头使用：邮箱、邮件、Email 等常见名称。",
            )
        )
    elif recipient_column not in excel_data.headers:
        issues.append(
            ValidationIssue(
                "error",
                f"收件人列「{recipient_column}」不存在于 Excel 表头中。",
            )
        )

    if cc_column and cc_column not in excel_data.headers:
        issues.append(
            ValidationIssue(
                "error",
                f"抄送列「{cc_column}」不存在于 Excel 表头中。",
            )
        )
    elif recipient_column and recipient_column in excel_data.headers:
        invalid_emails = find_invalid_recipient_emails(excel_data, recipient_column)
        if invalid_emails:
            preview_lines = [
                f"第 {row_index} 行: {email or '(空)'} ({reason})"
                for row_index, email, reason in invalid_emails[:20]
            ]
            if len(invalid_emails) > 20:
                preview_lines.append(f"... 另有 {len(invalid_emails) - 20} 条")
            issues.append(
                ValidationIssue(
                    "error",
                    "以下收件邮箱格式有误：\n" + "\n".join(preview_lines),
                )
            )

    placeholders = {match.group(1).strip() for match in PLACEHOLDER_PATTERN.finditer(normalize_placeholder_text(template_text))}
    missing_columns = sorted(
        column for column in placeholders if column and not _column_in_headers(column, excel_data.headers)
    )
    if missing_columns:
        issues.append(
            ValidationIssue(
                "warning",
                f"模板占位符在 Excel 中不存在：{', '.join(missing_columns)}",
            )
        )

    normalized_subject = normalize_placeholder_text(subject)
    subject_placeholders = {match.group(1).strip() for match in PLACEHOLDER_PATTERN.finditer(normalized_subject)}
    missing_subject_columns = sorted(
        column for column in subject_placeholders if column and not _column_in_headers(column, excel_data.headers)
    )
    if missing_subject_columns:
        issues.append(
            ValidationIssue(
                "warning",
                f"主题占位符在 Excel 中不存在：{', '.join(missing_subject_columns)}",
            )
        )

    if not attachment_paths:
        issues.append(ValidationIssue("warning", "未选择任何附件。"))
    else:
        for attachment_path in attachment_paths:
            path = Path(attachment_path)
            if not path.is_file():
                issues.append(ValidationIssue("error", f"附件不存在：{path}"))

    return issues


def find_email_column(headers: Iterable[str]) -> str | None:
    """在表头中查找收件邮箱列。"""
    header_list = [header for header in headers if header]
    exact = _find_column_by_names(header_list, EMAIL_COLUMN_NAMES)
    if exact:
        return exact

    cc_col = find_cc_column(header_list)
    for header in header_list:
        if header == cc_col:
            continue
        text = header.strip()
        lowered = text.lower()
        if "抄送" in text or lowered.startswith("cc"):
            continue
        if (
            "邮箱" in text
            or "email" in lowered
            or "e-mail" in lowered
            or text in {"邮件", "Mail", "mail"}
        ):
            return header
    return None


def find_cc_column(headers: Iterable[str]) -> str | None:
    """在表头中查找抄送邮箱列。"""
    return _find_column_by_names(headers, CC_COLUMN_NAMES)


def _find_column_by_names(headers: Iterable[str], names: tuple[str, ...]) -> str | None:
    header_list = list(headers)
    for name in names:
        if name in header_list:
            return name
    lowered = {header.lower(): header for header in header_list}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    return None


def is_valid_email(email: str) -> bool:
    text = email.strip()
    if not text:
        return False
    return EMAIL_FORMAT_PATTERN.fullmatch(text) is not None


def find_invalid_recipient_emails(
    excel_data: ExcelData,
    email_column: str,
) -> list[tuple[int, str, str]]:
    """返回格式无效的收件邮箱：(行号, 邮箱, 原因)。"""
    invalid: list[tuple[int, str, str]] = []
    for row_index, row_dict in enumerate(excel_data.rows, start=1):
        email = str(row_dict.get(email_column, "")).strip()
        if not email:
            invalid.append((row_index, "", "收件邮箱为空"))
        elif not is_valid_email(email):
            invalid.append((row_index, email, "邮箱格式不正确"))
    return invalid


def parse_email_list(value: Any) -> list[str]:
    """解析单元格中的多个邮箱，支持中英文逗号、分号分隔。"""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[,，;；\s]+", text)
    return [part.strip() for part in parts if part.strip()]


def convert_legacy_placeholders(template_text: str) -> str:
    """将旧格式【列名】转换为 {{列名}}。"""

    def replace(match: re.Match[str]) -> str:
        inner = match.group(1).strip()
        # 【{{列名}}】→ {{列名}}，避免生成 {{{{列名}}}}
        if inner.startswith("{{") and inner.endswith("}}"):
            column = inner[2:-2].strip()
            return f"{{{{{column}}}}}"
        return f"{{{{{inner}}}}}"

    return LEGACY_PLACEHOLDER_PATTERN.sub(replace, template_text)


def _column_in_headers(column: str, headers: Iterable[str]) -> bool:
    header_list = [header for header in headers if header]
    if column in header_list:
        return True
    stripped = column.strip()
    return any(header.strip() == stripped for header in header_list)


def normalize_placeholder_text(text: str) -> str:
    """统一占位符写法（全角括号、旧格式等）。"""
    text = text.replace("\uff5b", "{").replace("\uff5d", "}")
    return convert_legacy_placeholders(text)


def _lookup_row_value(row_dict: dict[str, Any], column: str) -> Any:
    if column in row_dict:
        return row_dict[column]
    stripped = column.strip()
    for key, value in row_dict.items():
        if key.strip() == stripped:
            return value
    return ""


def render_template(template_text: str, row_dict: dict[str, Any]) -> str:
    """用行数据替换模板中的 {{列名}} 占位符。"""

    def replace(match: re.Match[str]) -> str:
        column = match.group(1).strip()
        value = _lookup_row_value(row_dict, column)
        return "" if value is None else str(value)

    return PLACEHOLDER_PATTERN.sub(replace, normalize_placeholder_text(template_text))


def build_email_preview(
    excel_data: ExcelData,
    row_index: int,
    template_text: str,
    subject: str,
    attachment_paths: Iterable[str | Path],
    *,
    email_column: str | None = None,
    cc_column: str | None = None,
) -> EmailPreview:
    """生成指定行的邮件预览。"""
    if row_index < 1 or row_index > len(excel_data.rows):
        raise ValueError(f"行号无效：{row_index}，有效范围 1-{len(excel_data.rows)}")

    row_dict = excel_data.rows[row_index - 1]
    recipient_column = email_column or find_email_column(excel_data.headers)
    if not recipient_column:
        raise ValueError("未找到收件邮箱列")

    cc_col = cc_column or find_cc_column(excel_data.headers)
    to_email = str(row_dict.get(recipient_column, "")).strip()
    cc_emails = parse_email_list(row_dict.get(cc_col, "")) if cc_col else []
    body = render_template(template_text, row_dict)
    attachment_list = [str(Path(path)) for path in attachment_paths]

    return EmailPreview(
        row_index=row_index,
        to_email=to_email,
        cc_emails=cc_emails,
        subject=render_template(subject, row_dict),
        body=body,
        attachment_paths=attachment_list,
    )


def build_email(
    *,
    smtp_config: SmtpConfig,
    to_email: str,
    subject: str,
    body: str,
    attachment_paths: Iterable[str | Path] | None = None,
    cc_emails: Iterable[str] | None = None,
) -> MIMEMultipart:
    """构建带固定多附件的邮件对象。"""
    message = MIMEMultipart()
    message["From"] = (
        formataddr((smtp_config.from_name, smtp_config.from_email))
        if smtp_config.from_name
        else smtp_config.from_email
    )
    message["To"] = to_email
    cc_list = [email.strip() for email in (cc_emails or []) if email and str(email).strip()]
    if cc_list:
        message["Cc"] = ", ".join(cc_list)
    message["Subject"] = Header(subject, "utf-8")
    message.attach(MIMEText(body, "html", "utf-8"))

    for attachment_path in attachment_paths or []:
        path = Path(attachment_path)
        if not path.is_file():
            raise FileNotFoundError(f"附件不存在: {path}")
        with path.open("rb") as file_obj:
            part = MIMEApplication(file_obj.read(), Name=path.name)
        part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", path.name))
        message.attach(part)

    return message


def send_email(smtp_config: SmtpConfig, message: MIMEMultipart) -> None:
    """通过 SMTP 发送单封邮件。"""
    if smtp_config.use_ssl:
        server: smtplib.SMTP = smtplib.SMTP_SSL(smtp_config.host, smtp_config.port, timeout=60)
    else:
        server = smtplib.SMTP(smtp_config.host, smtp_config.port, timeout=60)

    try:
        if smtp_config.use_tls and not smtp_config.use_ssl:
            server.starttls()
        if smtp_config.user:
            server.login(smtp_config.user, smtp_config.password)
        server.send_message(message)
    finally:
        server.quit()


def send_batch(
    excel_data: ExcelData,
    template_text: str,
    subject: str,
    smtp_config: SmtpConfig,
    attachment_paths: Iterable[str | Path],
    *,
    email_column: str | None = None,
    cc_column: str | None = None,
    progress_callback: Callable[[int, int, SendResult], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
    should_stop: Callable[[], bool] | None = None,
) -> BatchSendSummary:
    """根据已加载的 Excel 数据批量发送邮件。"""
    if not smtp_config.host or not smtp_config.user or not smtp_config.password:
        raise ValueError("SMTP 配置不完整，请设置 host/user/password")

    attachment_list = [Path(path) for path in attachment_paths]
    for attachment in attachment_list:
        if not attachment.is_file():
            raise FileNotFoundError(f"附件不存在: {attachment}")

    recipient_column = email_column or find_email_column(excel_data.headers)
    if not recipient_column:
        raise ValueError("未找到收件邮箱列，请使用：邮箱、邮件、Email、email 之一")

    cc_col = cc_column or find_cc_column(excel_data.headers)

    rows = excel_data.rows
    interval = smtp_config.send_interval
    summary = BatchSendSummary(total=len(rows))

    for row_index, row_dict in enumerate(rows, start=1):
        if should_stop and should_stop():
            if log_callback:
                log_callback(f"发送任务已取消，停在第 {row_index} 行")
            logger.info("发送任务已取消，停在第 %s 行", row_index)
            break

        recipient = str(row_dict.get(recipient_column, "")).strip()
        cc_emails = parse_email_list(row_dict.get(cc_col, "")) if cc_col else []
        cc_text = ", ".join(cc_emails)
        if not recipient:
            summary.skipped_count += 1
            result = SendResult(
                row_index=row_index,
                recipient="",
                success=False,
                error=f"第 {row_index} 行缺少收件邮箱",
            )
            summary.results.append(result)
            message = f"跳过第 {row_index} 行：缺少收件邮箱"
            logger.warning(message)
            if log_callback:
                log_callback(message)
            if progress_callback:
                progress_callback(row_index, len(rows), result)
            continue

        try:
            body = render_template(template_text, row_dict)
            message = build_email(
                smtp_config=smtp_config,
                to_email=recipient,
                subject=render_template(subject, row_dict),
                body=body,
                attachment_paths=attachment_list,
                cc_emails=cc_emails,
            )
            send_email(smtp_config, message)
            result = SendResult(
                row_index=row_index,
                recipient=recipient,
                success=True,
                cc=cc_text,
            )
            summary.success_count += 1
            cc_log = f"，抄送 {cc_text}" if cc_text else ""
            message = f"发送成功：第 {row_index} 行 -> {recipient}{cc_log}"
            logger.info(message)
            if log_callback:
                log_callback(message)
        except Exception as exc:  # noqa: BLE001 - 批量发送需记录每行失败原因
            result = SendResult(
                row_index=row_index,
                recipient=recipient,
                success=False,
                error=str(exc),
                cc=cc_text,
            )
            summary.failure_count += 1
            message = f"发送失败：第 {row_index} 行 -> {recipient}，原因：{exc}"
            logger.error(message)
            if log_callback:
                log_callback(message)

        summary.results.append(result)
        if progress_callback:
            progress_callback(row_index, len(rows), result)

        if interval > 0 and row_index < len(rows):
            time.sleep(interval)

    return summary


def export_failed_send_results(
    excel_data: ExcelData,
    summary: BatchSendSummary,
    output_path: str | Path,
) -> Path:
    """将发送失败的记录导出到 Excel。"""
    from openpyxl import Workbook

    failed_results = [result for result in summary.results if not result.success]
    if not failed_results:
        raise ValueError("没有失败记录可导出")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "发送失败"

    headers = ["行号", "收件邮箱", "失败原因", *excel_data.headers]
    worksheet.append(headers)

    for result in failed_results:
        row_dict: dict[str, Any] = {}
        if 1 <= result.row_index <= len(excel_data.rows):
            row_dict = excel_data.rows[result.row_index - 1]
        row = [result.row_index, result.recipient, result.error]
        row.extend(row_dict.get(header, "") for header in excel_data.headers)
        worksheet.append(row)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def send_batch_from_file(
    *,
    excel_path: str | Path,
    template_text: str,
    subject: str,
    attachment_paths: Iterable[str | Path] | None = None,
    smtp_config: SmtpConfig | None = None,
    env_path: str | Path | None = None,
    send_interval: float | None = None,
    email_column: str | None = None,
    cc_column: str | None = None,
    on_progress: Callable[[SendResult], None] | None = None,
    should_stop: Callable[[], bool] | None = None,
) -> BatchSendSummary:
    """从 Excel 文件路径批量发送邮件。"""
    config = smtp_config or load_smtp_config(env_path)
    if send_interval is not None:
        config.send_interval = send_interval

    excel_data = load_excel(excel_path)

    def progress_callback(current: int, total: int, result: SendResult) -> None:
        if on_progress:
            on_progress(result)

    return send_batch(
        excel_data,
        template_text,
        subject,
        config,
        attachment_paths or [],
        email_column=email_column,
        cc_column=cc_column,
        progress_callback=progress_callback if on_progress else None,
        should_stop=should_stop,
    )


def load_template(path: str | Path, *, convert_legacy: bool = False) -> str:
    """读取模板文件，可选将旧占位符转换为 {{列名}}。"""
    text = Path(path).read_text(encoding="utf-8")
    if convert_legacy:
        return convert_legacy_placeholders(text)
    return text
